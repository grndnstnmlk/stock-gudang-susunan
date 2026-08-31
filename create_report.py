import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# ==============================================================================
# 1. LOAD MASTER DATABASE (DOKUMEN REKAP NO GUD, BARKOT, GRADE, KG)
# ==============================================================================
def is_valid_barcode(b):
    if b is None:
        return False
    s = str(b).strip()
    # Barcode harus memiliki digit/angka dan bukan status OUT/None/-/?
    if not any(c.isdigit() for c in s):
        return False
    if s.lower() in ['out', 'none', '-', '', 'null']:
        return False
    return True

def load_master_database(master_filepath='Dokumen_Rekap_NoGud_Barkot_Kg.xlsx'):
    master_dict = {}
    if not os.path.exists(master_filepath):
        print(f"Warning: {master_filepath} not found. Running with empty master dictionary.")
        return master_dict

    wb_master = openpyxl.load_workbook(master_filepath, data_only=True)
    
    def update_entry(no_gud, barkot, grade, kg, status, ket):
        if no_gud is None or str(no_gud).strip() == '':
            return
        k = str(no_gud).strip()
        has_new_barkot = is_valid_barcode(barkot)
        b_str = str(barkot).strip() if has_new_barkot else ''
        kg_val = '' if kg in [None, '', '-'] else kg
        grade_str = '' if grade in [None, '', '-'] else str(grade).strip()
        status_str = '' if status in [None, '', '-'] else str(status).strip()
        ket_str = '' if ket in [None, '', '-'] else str(ket).strip()

        if k not in master_dict:
            master_dict[k] = {
                'no_gud': no_gud,
                'barkot': b_str,
                'grade': grade_str,
                'kg': kg_val,
                'status': status_str,
                'ket': ket_str
            }
        else:
            has_old_barkot = is_valid_barcode(master_dict[k].get('barkot'))
            # Aturan:
            # 1. Jika data baru (tanggal selanjutnya) memiliki barcode valid -> otomatis timpa data lama.
            # 2. Jika data baru TIDAK ada barcode (misal OUT/kosong), tapi data lama SUDAH punya barcode -> pertahankan barcode lama!
            # 3. Jika dua-duanya belum punya barcode -> perbarui atribut status/grade/kg jika ada.
            if has_new_barkot:
                master_dict[k] = {
                    'no_gud': no_gud,
                    'barkot': b_str,
                    'grade': grade_str or master_dict[k]['grade'],
                    'kg': kg_val if kg_val != '' else master_dict[k]['kg'],
                    'status': 'SELESAI' if status_str != 'OUT' else master_dict[k]['status'],
                    'ket': ket_str or master_dict[k]['ket']
                }
            elif not has_old_barkot:
                master_dict[k] = {
                    'no_gud': no_gud,
                    'barkot': '',
                    'grade': grade_str or master_dict[k]['grade'],
                    'kg': kg_val if kg_val != '' else master_dict[k]['kg'],
                    'status': status_str or master_dict[k]['status'],
                    'ket': ket_str or master_dict[k]['ket']
                }

    # 1. Baca dari sheet harian (21 s/d 30 Agustus)
    daily_sheets = [s for s in wb_master.sheetnames if s not in ['Semua Data (Master)', 'Ringkasan Per Tanggal']]
    for sname in daily_sheets:
        ws = wb_master[sname]
        for r in range(5, ws.max_row + 1):
            update_entry(ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value)

    # 2. Baca / lengkapi dari sheet Semua Data (Master)
    if 'Semua Data (Master)' in wb_master.sheetnames:
        ws_m = wb_master['Semua Data (Master)']
        for r in range(5, ws_m.max_row + 1):
            update_entry(ws_m.cell(r, 3).value, ws_m.cell(r, 4).value, ws_m.cell(r, 5).value, ws_m.cell(r, 6).value, ws_m.cell(r, 7).value, ws_m.cell(r, 8).value)

    # 3. Custom / User Overrides
    custom_overrides = {
        '259': {'no_gud': 259, 'barkot': '30164', 'grade': '55', 'kg': 55, 'status': 'SELESAI', 'ket': ''},
    }
    for k, v in custom_overrides.items():
        master_dict[k] = v

    return master_dict

master_bal_dict = load_master_database('Dokumen_Rekap_NoGud_Barkot_Kg.xlsx')

def lookup_bal_info(val):
    if val is None or val == '':
        return '', '', '', ''
    s = str(val).strip()
    if s in master_bal_dict:
        m = master_bal_dict[s]
        return val, m['barkot'], m['kg'], m['grade']
    s_clean = s.lstrip('0')
    if s_clean and s_clean in master_bal_dict:
        m = master_bal_dict[s_clean]
        return val, m['barkot'], m['kg'], m['grade']
    if '/' in s:
        parts = [p.strip() for p in s.split('/')]
        barkots = []
        kgs = []
        grades = []
        for p in parts:
            p_clean = p.lstrip('0') if p.lstrip('0') in master_bal_dict else p
            if p in master_bal_dict:
                m = master_bal_dict[p]
            elif p_clean in master_bal_dict:
                m = master_bal_dict[p_clean]
            else:
                m = None
            if m:
                if m['barkot']: barkots.append(m['barkot'])
                if m['kg'] != '': kgs.append(str(m['kg']))
                if m['grade']: grades.append(m['grade'])
        return val, '/'.join(barkots) if barkots else '-', '/'.join(kgs) if kgs else '-', '/'.join(grades) if grades else '-'
    return val, '-', '-', '-'

# ==============================================================================
# 2. DATA SUSUNAN BAL GUDANG (16 BLOK)
# ==============================================================================
blocks_data = {
    1: {
        "title": "BLOK 01",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [305, 18, 133, 144, 234, 274], # T4 (Atas)
            [508, 214, 460, 462, 110, 123], # T3
            [283, 23, 467, 509, 177, 193],  # T2
            [494, 497, 453, 474, 116, 310], # T1 (Dasar)
        ]
    },
    2: {
        "title": "BLOK 02",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [125, 444, 158, 448, 108, 112], # T4 (Atas)
            [493, 439, 438, 98, 171, 124],  # T3
            [183, 368, 160, 491, 469, 504], # T2
            [164, 288, 15, 227, 416, 99],   # T1 (Dasar)
        ]
    },
    3: {
        "title": "BLOK 03",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [230, 199, 503, 495, 32, 437],  # T4 (Atas)
            [180, 473, 432, 456, 282, 91],  # T3
            [222, 217, 502, 440, 487, 107], # T2
            [408, 465, 296, 421, 463, 206], # T1 (Dasar)
        ]
    },
    4: {
        "title": "BLOK 04",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [100, 492, 226, 506, 374, 370], # T4 (Atas)
            [500, 436, 475, 201, 399, 111], # T3
            [498, 501, 130, 489, 115, 19],  # T2
            [466, 447, 431, 277, 496, 468], # T1 (Dasar)
        ]
    },
    5: {
        "title": "BLOK 05",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [412, 430, 451, 490, 109, 4],   # T4 (Atas)
            [257, 386, 246, 241, 313, 260], # T3
            [381, 224, 397, 245, 413, 272], # T2
            [499, 210, 309, 373, 332, 405], # T1 (Dasar)
        ]
    },
    6: {
        "title": "BLOK 06",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5"],
        "data": [
            [388, 367, 349, 336, 300], # T4 (Atas)
            [472, 334, 315, 387, 103], # T3
            [423, 389, 365, 263, 458], # T2
            [211, 258, 433, 398, 122], # T1 (Dasar)
        ]
    },
    7: {
        "title": "BLOK 07",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4"],
        "data": [
            [318, 427, 395, 311], # T4 (Atas)
            [380, 350, 330, 329], # T3
            [377, 326, 339, 375], # T2
            [320, 384, 371, 352], # T1 (Dasar)
        ]
    },
    8: {
        "title": "BLOK 08",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4"],
        "data": [
            [267, 271, 90, 341],  # T4 (Atas)
            [340, 382, 333, 265], # T3
            [425, 369, 383, 393], # T2
            [325, 379, 324, 385], # T1 (Dasar)
        ]
    },
    9: {
        "title": "BLOK 09",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4"],
        "data": [
            [418, 426, 337, 323], # T4 (Atas)
            [419, 424, 254, 378], # T3
            [275, 197, 331, 396], # T2
            [321, 415, 434, 278], # T1 (Dasar)
        ]
    },
    10: {
        "title": "BLOK 10",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4"],
        "data": [
            [488, 344, 428, 470], # T4 (Atas)
            [420, 358, 50, 406],  # T3
            [400, 390, 328, 366], # T2
            [421, 351, 255, 392], # T1 (Dasar)
        ]
    },
    11: {
        "title": "BLOK 11",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5"],
        "data": [
            ["",  247, 414, 346, 338], # T4 (Atas)
            [565, 154, 505, 259, 228], # T3
            [567, 422, 347, 327, 355], # T2
            [534, 289, 410, 357, 404], # T1 (Dasar)
        ]
    },
    12: {
        "title": "BLOK 12",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [517, 208, 342, 169, 216, 459], # T4 (Atas)
            [519, 529, 445, 403, 361, 455], # T3
            [30,  528, 312, 316, 407, 231], # T2
            [512, 545, 394, 322, 335, 264], # T1 (Dasar)
        ]
    },
    13: {
        "title": "BLOK 13",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [535, 531, 63,  168, 372, 402], # T4 (Atas)
            [527, 533, 233, 461, 446, 132], # T3
            [126, 16,  140, 141, 178, 401], # T2
            ["08", 541, 173, 409, 442, 429], # T1 (Dasar)
        ]
    },
    14: {
        "title": "BLOK 14",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [532, 571, 273, 449, 441, 181], # T4 (Atas)
            [482, 13,  219, 232, 454, 457], # T3
            [526, 308, 49,  411, 319, 452], # T2
            [522, 543, 450, 256, 376, 175], # T1 (Dasar)
        ]
    },
    15: {
        "title": "BLOK 15",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [525, 537, 276, 564, 149, 443], # T4 (Atas)
            [521, 536, 575, 223, 220, 435], # T3
            [481, 566, 559, 572, 356, 285], # T2
            ["03", 553, 179, 574, 150, 343], # T1 (Dasar)
        ]
    },
    16: {
        "title": "BLOK 16",
        "headers": ["Saf 1", "Saf 2", "Saf 3", "Saf 4", "Saf 5", "Saf 6"],
        "data": [
            [113, 552, 555, 558, 549, 307], # T4 (Atas)
            [480, 548, 542, 557, 544, 561], # T3
            [523, 539, 547, 554, 540, 538], # T2
            [524, 551, 546, 573, 64, 560],  # T1 (Dasar)
        ]
    },
}

all_tingkat_info = [
    ('T7 (Atas)', True),
    ('T6', True),
    ('T5', True),
    ('T4', False),
    ('T3', False),
    ('T2', False),
    ('T1 (Dasar)', False)
]

wb = openpyxl.Workbook()

# ==============================================================================
# 3. STYLES DEFINITION (HEMAT TINTA, FONT TEGAS, BORDER RAPI)
# ==============================================================================
font_main_title = Font(name='Segoe UI', size=16, bold=True, color='000000')
font_subtitle = Font(name='Segoe UI', size=11, bold=True, color='333333')
font_date = Font(name='Segoe UI', size=10, italic=True, color='333333')

font_block_hdr = Font(name='Segoe UI', size=12, bold=True, color='000000')
font_saf_super = Font(name='Segoe UI', size=11, bold=True, color='000000')
font_subhdr_tiny = Font(name='Segoe UI', size=9, bold=True, color='333333')
font_subhdr = Font(name='Segoe UI', size=11, bold=True, color='000000')

def format_kg_display(kg_val):
    if kg_val is None or kg_val == '':
        return ''
    if kg_val == '-':
        return '-'
    s = str(kg_val).strip()
    if '/' in s:
        parts = [p.strip() for p in s.split('/')]
        return '/'.join([f"{p} kg" for p in parts])
    return f"{s} kg"

font_data_nogud = Font(name='Segoe UI', size=11, bold=True, color='0F172A')
font_data_barkot = Font(name='Segoe UI', size=10, bold=False, color='334155')
font_data_kg = Font(name='Segoe UI', size=10, bold=True, color='0F766E')
font_data_jumbo = Font(name='Segoe UI', size=14, bold=True, color='000000')
font_tingkat = Font(name='Segoe UI', size=11, bold=True, color='000000')
font_data_regular = Font(name='Segoe UI', size=11, bold=False, color='000000')

fill_banner_eco = PatternFill(start_color='E8EFF5', end_color='E8EFF5', fill_type='solid')
fill_subhdr_eco = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
fill_tingkat_eco = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
fill_kg_col = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Border halus & rapi (slim, bersih, tidak tebal/bold saat print)
border_side_clean = Side(border_style='thin', color='BFBFBF')
border_all_black = Border(left=border_side_clean, right=border_side_clean, top=border_side_clean, bottom=border_side_clean)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

def configure_a4_print(ws, orientation='portrait', fit_w=1, fit_h=0):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = fit_h
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True
    ws.print_options.gridLines = False

# ==============================================================================
# SHEET 1: STOCK LENGKAP (NO GUDANG + BARKOT + KG DI SEBELAH KANAN)
# ==============================================================================
ws_full = wb.active
ws_full.title = 'Stock Lengkap (NoGud+Barkot+Kg)'
ws_full.views.sheetView[0].showGridLines = True
configure_a4_print(ws_full, orientation='landscape', fit_w=1, fit_h=0)

# Title & Date
ws_full['B2'] = 'STOCK GUDANG DJARUM'
ws_full['B2'].font = font_main_title
ws_full['B2'].alignment = align_left
ws_full.row_dimensions[2].height = 20

ws_full['B3'] = 'LAPORAN SUSUNAN BAL GUDANG (NO GUDANG, NOMOR BARKOT & BERAT KG)'
ws_full['B3'].font = font_subtitle
ws_full['B3'].alignment = align_left
ws_full.row_dimensions[3].height = 16

ws_full['B4'] = 'Minggu 30/8/2026'
ws_full['B4'].font = font_date
ws_full['B4'].alignment = align_left
ws_full.row_dimensions[4].height = 14

curr_r = 6
for b_idx in range(1, 17):
    b = blocks_data[b_idx]
    num_saf = len(b["headers"])
    # 1 column for Tingkat + 3 columns per saf (No Gud, Barkot, Kg)
    total_cols = 1 + (num_saf * 3)
    start_col = 2
    end_col = start_col + total_cols - 1
    
    # 1. Block Header Banner dengan Keterangan Arah
    banner_text = f"{b['title']} (Saf 1 Utara - Saf {num_saf} Selatan)"
    ws_full.merge_cells(start_row=curr_r, start_column=start_col, end_row=curr_r, end_column=end_col)
    hdr_cell = ws_full.cell(curr_r, start_col, banner_text)
    hdr_cell.font = font_block_hdr
    hdr_cell.alignment = align_center
    for c in range(start_col, end_col + 1):
        ws_full.cell(curr_r, c).fill = fill_banner_eco
        ws_full.cell(curr_r, c).border = border_all_black
    ws_full.row_dimensions[curr_r].height = 20
    curr_r += 1
    
    # 2. Table Column Headers:
    # Row 1: Tingkat (merged 2 rows), Saf 1 (merged 3 cols), Saf 2 (merged 3 cols), ...
    ws_full.merge_cells(start_row=curr_r, start_column=start_col, end_row=curr_r + 1, end_column=start_col)
    t_hdr = ws_full.cell(curr_r, start_col, 'Tingkat')
    t_hdr.font = font_subhdr
    t_hdr.fill = fill_subhdr_eco
    t_hdr.alignment = align_center
    t_hdr.border = border_all_black
    ws_full.cell(curr_r + 1, start_col).border = border_all_black
    ws_full.cell(curr_r + 1, start_col).fill = fill_subhdr_eco
    
    for s_idx, saf_title in enumerate(b["headers"]):
        col_s = start_col + 1 + (s_idx * 3)
        col_e = col_s + 2
        ws_full.merge_cells(start_row=curr_r, start_column=col_s, end_row=curr_r, end_column=col_e)
        s_cell = ws_full.cell(curr_r, col_s, saf_title)
        s_cell.font = font_saf_super
        s_cell.alignment = align_center
        for c in range(col_s, col_e + 1):
            ws_full.cell(curr_r, c).fill = fill_subhdr_eco
            ws_full.cell(curr_r, c).border = border_all_black
            
        # Row 2 Sub-headers: No Gud | Barkot | Kg
        sub_names = ['No Gud', 'Barkot', 'Kg']
        for sub_idx, sname in enumerate(sub_names):
            sub_col = col_s + sub_idx
            sub_cell = ws_full.cell(curr_r + 1, sub_col, sname)
            sub_cell.font = font_subhdr_tiny
            sub_cell.fill = fill_subhdr_eco
            sub_cell.alignment = align_center
            sub_cell.border = border_all_black
            
    ws_full.row_dimensions[curr_r].height = 17
    ws_full.row_dimensions[curr_r + 1].height = 15
    curr_r += 2
    
    # 3. Data Rows (7 Tingkat)
    data_idx = 0
    for t_name, is_empty in all_tingkat_info:
        # Tingkat label
        t_cell = ws_full.cell(curr_r, start_col, t_name)
        t_cell.font = font_tingkat
        t_cell.fill = fill_tingkat_eco
        t_cell.alignment = align_center
        t_cell.border = border_all_black
        
        if is_empty:
            for s_idx in range(num_saf):
                col_s = start_col + 1 + (s_idx * 3)
                for sub_idx in range(3):
                    c_cell = ws_full.cell(curr_r, col_s + sub_idx, '')
                    c_cell.fill = fill_white
                    c_cell.border = border_all_black
        else:
            row_vals = b["data"][data_idx]
            data_idx += 1
            for s_idx, val in enumerate(row_vals):
                no_gud_val, barkot_val, kg_val, grade_val = lookup_bal_info(val)
                col_s = start_col + 1 + (s_idx * 3)
                
                # Col 1: No Gud
                c1 = ws_full.cell(curr_r, col_s, no_gud_val)
                c1.font = font_data_nogud
                c1.alignment = align_center
                c1.fill = fill_white
                c1.border = border_all_black
                
                # Col 2: Barkot
                c2 = ws_full.cell(curr_r, col_s + 1, barkot_val)
                c2.font = font_data_barkot
                c2.alignment = align_center
                c2.fill = fill_white
                c2.border = border_all_black
                
                # Col 3: Kg (Format pembeda dengan satuan 'kg' & subtle background)
                c3 = ws_full.cell(curr_r, col_s + 2, format_kg_display(kg_val))
                c3.font = font_data_kg
                c3.alignment = align_center
                c3.fill = fill_kg_col
                c3.border = border_all_black
                
        ws_full.row_dimensions[curr_r].height = 17
        curr_r += 1
        
    curr_r += 1
    ws_full.row_dimensions[curr_r].height = 5 # Spacer
    
    # Page breaks: Page 1 memuat 2 blok (karena ada judul utama), Page 2 dst memuat 3 blok
    if b_idx in [2, 5, 8, 11, 14]:
        ws_full.row_breaks.append(Break(id=curr_r))
        
    curr_r += 1

# Column widths for Sheet 1
ws_full.column_dimensions['A'].width = 2
ws_full.column_dimensions['B'].width = 11 # Tingkat
for col_idx in range(3, 22):
    letter = get_column_letter(col_idx)
    sub_type = (col_idx - 3) % 3
    if sub_type == 0:
        ws_full.column_dimensions[letter].width = 8.5  # No Gud
    elif sub_type == 1:
        ws_full.column_dimensions[letter].width = 9.5 # Barkot
    else:
        ws_full.column_dimensions[letter].width = 7.5  # Kg

# ==============================================================================
# SHEET 2: STOCK RINGKAS (NO GUD SAJA - CETAK PORTRAIT 4 BLOK PER HALAMAN)
# ==============================================================================
ws_portrait = wb.create_sheet(title='Stock Ringkas (No Gud Saja)')
ws_portrait.views.sheetView[0].showGridLines = True
configure_a4_print(ws_portrait, orientation='portrait', fit_w=1, fit_h=4)

ws_portrait['B2'] = 'STOCK GUDANG DJARUM'
ws_portrait['B2'].font = font_main_title
ws_portrait['B2'].alignment = align_left
ws_portrait.row_dimensions[2].height = 20

ws_portrait['B3'] = 'Minggu 30/8/2026'
ws_portrait['B3'].font = font_subtitle
ws_portrait['B3'].alignment = align_left
ws_portrait.row_dimensions[3].height = 16

curr_r = 5
for b_idx in range(1, 17):
    b = blocks_data[b_idx]
    num_cols = len(b["headers"])
    end_col = 2 + num_cols
    
    # Block Header Banner (BLOK 01 (Saf 1 Utara - Saf 6 Selatan))
    banner_text = f"{b['title']} (Saf 1 Utara - Saf {num_cols} Selatan)"
    ws_portrait.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r, end_column=end_col)
    hdr_cell = ws_portrait.cell(curr_r, 2, banner_text)
    hdr_cell.font = font_block_hdr
    hdr_cell.alignment = align_center
    for c in range(2, end_col + 1):
        ws_portrait.cell(curr_r, c).fill = fill_banner_eco
        ws_portrait.cell(curr_r, c).border = border_all_black
    ws_portrait.row_dimensions[curr_r].height = 21
    
    # Table Column Headers (Tingkat, Saf 1, Saf 2, ...)
    curr_r += 1
    t_hdr = ws_portrait.cell(curr_r, 2, 'Tingkat')
    t_hdr.font = font_subhdr
    t_hdr.fill = fill_subhdr_eco
    t_hdr.alignment = align_center
    t_hdr.border = border_all_black
    
    for c_idx, h_name in enumerate(b["headers"]):
        col_num = 3 + c_idx
        c_cell = ws_portrait.cell(curr_r, col_num, h_name)
        c_cell.font = font_subhdr
        c_cell.fill = fill_subhdr_eco
        c_cell.alignment = align_center
        c_cell.border = border_all_black
    ws_portrait.row_dimensions[curr_r].height = 20
    
    # Data Rows (7 Tiers)
    data_idx = 0
    for t_name, is_empty in all_tingkat_info:
        curr_r += 1
        t_cell = ws_portrait.cell(curr_r, 2, t_name)
        t_cell.font = font_tingkat
        t_cell.fill = fill_tingkat_eco
        t_cell.alignment = align_center
        t_cell.border = border_all_black
        
        if is_empty:
            for c_idx in range(num_cols):
                col_num = 3 + c_idx
                d_cell = ws_portrait.cell(curr_r, col_num, "")
                d_cell.fill = fill_white
                d_cell.border = border_all_black
        else:
            row_vals = b["data"][data_idx]
            data_idx += 1
            for c_idx, val in enumerate(row_vals):
                col_num = 3 + c_idx
                d_cell = ws_portrait.cell(curr_r, col_num, val)
                d_cell.font = font_data_jumbo
                d_cell.fill = fill_white
                d_cell.alignment = align_center
                d_cell.border = border_all_black
        ws_portrait.row_dimensions[curr_r].height = 18
        
    curr_r += 1
    ws_portrait.row_dimensions[curr_r].height = 6 # Spacer
    
    # Page Break tepat setelah Blok 4, Blok 8, Blok 12 (Tepat 4 blok per halaman)
    if b_idx in [4, 8, 12]:
        ws_portrait.row_breaks.append(Break(id=curr_r))
        
    curr_r += 1

# Column widths for Portrait Sheet
ws_portrait.column_dimensions['A'].width = 2
ws_portrait.column_dimensions['B'].width = 12.0
for c in ['C', 'D', 'E', 'F', 'G', 'H']:
    ws_portrait.column_dimensions[c].width = 11.5

# ==============================================================================
# SHEET 3: FORMAT GRID (c.xlsx) - DENGAN NO GUD, BARKOT, DAN KG
# ==============================================================================
ws2 = wb.create_sheet(title='Format Grid (c.xlsx)')
ws2.views.sheetView[0].showGridLines = True
configure_a4_print(ws2, orientation='landscape', fit_w=1, fit_h=4)

ws2['A2'] = 'STOCK GUDANG DJARUM'
ws2['A2'].font = font_main_title
ws2['A3'] = 'Minggu 30/8/2026'
ws2['A3'].font = font_subtitle

grid_start_r = 5
for b_num in range(1, 17):
    matrix = blocks_data[b_num]["data"]
    num_saf = len(blocks_data[b_num]["headers"])
    total_grid_cols = num_saf * 3
    
    # 3 Baris Kosong di Bagian Atas (Tingkat 7, Tingkat 6, Tingkat 5)
    for r_empty in range(3):
        for c_idx in range(total_grid_cols):
            cell = ws2.cell(row=grid_start_r + r_empty, column=c_idx + 1, value=None)
            cell.border = border_all_black
            
    # 4 Baris Data Terisi di Bagian Bawah (Tingkat 4, Tingkat 3, Tingkat 2, Tingkat 1)
    for r_idx, row_data in enumerate(matrix):
        for s_idx, val in enumerate(row_data):
            no_gud_val, barkot_val, kg_val, grade_val = lookup_bal_info(val)
            c_base = (s_idx * 3) + 1
            
            # No Gud
            c1 = ws2.cell(row=grid_start_r + 3 + r_idx, column=c_base, value=no_gud_val)
            c1.font = font_data_nogud
            c1.alignment = align_center
            c1.border = border_all_black
            
            # Barkot
            c2 = ws2.cell(row=grid_start_r + 3 + r_idx, column=c_base + 1, value=barkot_val)
            c2.font = font_data_barkot
            c2.alignment = align_center
            c2.border = border_all_black
            
            # Kg
            c3 = ws2.cell(row=grid_start_r + 3 + r_idx, column=c_base + 2, value=format_kg_display(kg_val))
            c3.font = font_data_kg
            c3.alignment = align_center
            c3.fill = fill_kg_col
            c3.border = border_all_black
            
    # Label Blok di sisi kanan
    b_col = total_grid_cols + 2
    b_cell = ws2.cell(row=grid_start_r + 3, column=b_col, value=b_num)
    b_cell.font = font_main_title
    b_cell.alignment = align_center
    
    if b_num in [4, 8, 12]:
        ws2.row_breaks.append(Break(id=grid_start_r + 6))
        
    grid_start_r += 7

for col in range(1, 22):
    col_letter = get_column_letter(col)
    sub_type = (col - 1) % 3
    if sub_type == 0:
        ws2.column_dimensions[col_letter].width = 8.5  # No Gud
    elif sub_type == 1:
        ws2.column_dimensions[col_letter].width = 9.5 # Barkot
    else:
        ws2.column_dimensions[col_letter].width = 7.5  # Kg

# ==============================================================================
# SHEET 4: REKAPITULASI & PENCARIAN BAL (Database Lengkap)
# ==============================================================================
ws3 = wb.create_sheet(title='Rekapitulasi & Cari Bal')
ws3.views.sheetView[0].showGridLines = True
configure_a4_print(ws3, orientation='landscape', fit_w=1, fit_h=0)

ws3['B2'] = 'STOCK GUDANG DJARUM'
ws3['B2'].font = font_main_title
ws3['B3'] = 'REKAPITULASI STOCK & DATABASE PENCARIAN BAL'
ws3['B3'].font = font_subtitle
ws3['B4'] = 'Minggu 30/8/2026'
ws3['B4'].font = font_date

# Table 1: Rekapitulasi per Blok
ws3['B6'] = 'Tabel 1: Rekapitulasi Stock per Blok (Kapasitas 7 Tingkat)'
ws3['B6'].font = font_subhdr

rekap_headers = ['No Blok', 'Nama Blok', 'Arah Saf', 'Jumlah Saf', 'Kapasitas (7 Tkt)', 'Terisi (Bal)', 'Kosong (Slot)', 'Total Berat (Kg)']
for c_idx, h in enumerate(rekap_headers):
    col_num = 2 + c_idx
    c_cell = ws3.cell(7, col_num, h)
    c_cell.font = font_block_hdr
    c_cell.fill = fill_banner_eco
    c_cell.alignment = align_center
    c_cell.border = border_all_black

ws3.row_dimensions[7].height = 24

r_row = 8
for b_num in range(1, 17):
    b = blocks_data[b_num]
    num_saf = len(b["headers"])
    kap_blok = num_saf * 7
    terisi_blok = sum(1 for row in b["data"] for val in row if val not in [None, ''])
    kosong_blok = kap_blok - terisi_blok
    arah_text = f"Saf 1 Utara - Saf {num_saf} Selatan"
    
    # Hitung total kg blok
    total_kg_blok = 0
    for row in b["data"]:
        for val in row:
            no, barkot, kg, grade = lookup_bal_info(val)
            if isinstance(kg, (int, float)):
                total_kg_blok += kg
            elif str(kg).isdigit():
                total_kg_blok += int(kg)
                
    ws3.cell(r_row, 2, b_num).alignment = align_center
    ws3.cell(r_row, 3, b["title"]).alignment = align_left
    ws3.cell(r_row, 4, arah_text).alignment = align_center
    ws3.cell(r_row, 5, num_saf).alignment = align_center
    ws3.cell(r_row, 6, kap_blok).alignment = align_center
    ws3.cell(r_row, 7, terisi_blok).alignment = align_center
    ws3.cell(r_row, 8, kosong_blok).alignment = align_center
    ws3.cell(r_row, 9, total_kg_blok).alignment = align_center
    
    for c in range(2, 10):
        cell = ws3.cell(r_row, c)
        cell.font = font_data_regular
        cell.border = border_all_black
        if b_num % 2 == 0:
            cell.fill = fill_tingkat_eco
    ws3.row_dimensions[r_row].height = 20
    r_row += 1

# Total Row
ws3.cell(r_row, 2, 'TOTAL').alignment = align_center
ws3.cell(r_row, 2).font = font_subhdr
ws3.merge_cells(start_row=r_row, start_column=2, end_row=r_row, end_column=4)
ws3.cell(r_row, 5, f'=SUM(E8:E{r_row-1})').alignment = align_center
ws3.cell(r_row, 6, f'=SUM(F8:F{r_row-1})').alignment = align_center
ws3.cell(r_row, 7, f'=SUM(G8:G{r_row-1})').alignment = align_center
ws3.cell(r_row, 8, f'=SUM(H8:H{r_row-1})').alignment = align_center
ws3.cell(r_row, 9, f'=SUM(I8:I{r_row-1})').alignment = align_center
for c in range(2, 10):
    cell = ws3.cell(r_row, c)
    cell.font = font_subhdr
    cell.fill = fill_subhdr_eco
    cell.border = border_all_black
ws3.row_dimensions[r_row].height = 22

# Table 2: Database Flat Table for Search
r_row += 3
ws3.cell(r_row, 2, 'Tabel 2: Database Seluruh Bal (Gunakan Filter Excel untuk Cari Nomor Bal)').font = font_subhdr
r_row += 1
db_start_r = r_row

db_headers = ['No', 'No Gudang', 'Nomor Barkot', 'Grade', 'Berat (Kg)', 'Blok', 'Tingkat', 'Saf / Kolom', 'Status Bal']
for c_idx, h in enumerate(db_headers):
    col_num = 2 + c_idx
    c_cell = ws3.cell(db_start_r, col_num, h)
    c_cell.font = font_block_hdr
    c_cell.fill = fill_banner_eco
    c_cell.alignment = align_center
    c_cell.border = border_all_black
ws3.row_dimensions[db_start_r].height = 22

item_idx = 1
for b_num in range(1, 17):
    b = blocks_data[b_num]
    tier_order = [
        (4, 'Tingkat 4'),
        (3, 'Tingkat 3'),
        (2, 'Tingkat 2'),
        (1, 'Tingkat 1 (Dasar)')
    ]
    for t_row_idx, (t_num, t_desc) in enumerate(tier_order):
        row_vals = b["data"][t_row_idx]
        for c_idx, val in enumerate(row_vals):
            if val in [None, '']:
                continue
            r_row += 1
            saf_name = b["headers"][c_idx]
            no_gud_val, barkot_val, kg_val, grade_val = lookup_bal_info(val)
            m_status = master_bal_dict.get(str(val).strip(), {}).get('status', 'SELESAI' if barkot_val not in ['-', ''] else '-')
            
            ws3.cell(r_row, 2, item_idx).alignment = align_center
            ws3.cell(r_row, 3, no_gud_val).alignment = align_center
            ws3.cell(r_row, 3).font = font_data_nogud
            ws3.cell(r_row, 4, barkot_val).alignment = align_center
            ws3.cell(r_row, 4).font = font_data_barkot
            ws3.cell(r_row, 5, grade_val).alignment = align_center
            ws3.cell(r_row, 6, format_kg_display(kg_val)).alignment = align_center
            ws3.cell(r_row, 6).font = font_data_kg
            ws3.cell(r_row, 7, b["title"]).alignment = align_center
            ws3.cell(r_row, 8, f'T{t_num}').alignment = align_center
            ws3.cell(r_row, 9, saf_name).alignment = align_center
            ws3.cell(r_row, 10, m_status).alignment = align_center
            
            for c in range(2, 11):
                cell = ws3.cell(r_row, c)
                cell.border = border_all_black
                if item_idx % 2 == 0:
                    cell.fill = fill_tingkat_eco
            ws3.row_dimensions[r_row].height = 20
            item_idx += 1

# Enable AutoFilter on Table 2
ws3.auto_filter.ref = f'B{db_start_r}:J{r_row}'

# Column widths for Sheet 4
ws3.column_dimensions['A'].width = 2.5
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 14
ws3.column_dimensions['J'].width = 14

# ==============================================================================
# 4. SAVE OUTPUT SPREADSHEETS
# ==============================================================================
target_files = [
    'Stock_Susunan_Bal_30_Agustus_2026.xlsx',
    'c.xlsx',
    r'C:\Users\xenov\Downloads\stock gudang susunan\c.xlsx',
    r'C:\Users\xenov\Downloads\stock gudang susunan\Stock_Susunan_Bal_30_Agustus_2026.xlsx'
]

for target in target_files:
    try:
        wb.save(target)
        print(f'Successfully saved to {target}')
    except Exception as e:
        print(f'Note: Could not save to {target}: {e}')

print('Report generation with Barkot, Kg, and Block Arah completed successfully!')
