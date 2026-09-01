import os
import json
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = "https://jrpklibocgicubevyshm.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImpycGtsaWJvY2dpY3ViZXZ5c2htIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODc4NTA3NjUsImV4cCI6MjEwMzQyNjc2NX0.xGoel8SNa2v9DcZBYwKcmjzGF7j6LJ-OQkr919JyYSc"

def fetch_supabase_records():
    """Mengambil seluruh rekaman barkot_data dari cloud Supabase (cekbarkot)."""
    url = f"{SUPABASE_URL}/rest/v1/barkot_data?select=*&order=tanggal.asc,no_gud.asc"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {SUPABASE_ANON_KEY}"
    }
    req = urllib.request.Request(url, headers=headers)
    print(f"[*] Menghubungi Supabase Cloud ({SUPABASE_URL})...")
    with urllib.request.urlopen(req, timeout=10) as response:
        records = json.loads(response.read().decode("utf-8"))
    print(f"[OK] Berhasil mengambil {len(records)} data bal dari Cek Barkot Supabase!")
    return records

def build_master_dict(records):
    """Membangun master dictionary dari record Supabase secara kronologis."""
    master_dict = {}
    for r in records:
        no_gud = r.get("no_gud")
        if not no_gud:
            continue
        k = str(no_gud).strip()
        barkot = str(r.get("barkot") or "").strip()
        grade = str(r.get("grade") or "").strip()
        kg = r.get("kg") if r.get("kg") is not None else ""
        is_done = r.get("is_done", False)
        
        has_barkot = barkot != "" and barkot.lower() not in ["none", "null", "-", "out"]
        status = "SELESAI" if has_barkot else ("SELESAI" if is_done else "-")
        
        if k not in master_dict or has_barkot:
            master_dict[k] = {
                "no_gud": no_gud,
                "barkot": barkot if has_barkot else "",
                "grade": grade or (master_dict[k]["grade"] if k in master_dict else ""),
                "kg": kg if kg != "" else (master_dict[k]["kg"] if k in master_dict else ""),
                "status": status,
                "ket": r.get("tanggal", "")
            }
        else:
            # Update atribut grade/kg jika belum terisi
            if grade and not master_dict[k]["grade"]:
                master_dict[k]["grade"] = grade
            if kg != "" and master_dict[k]["kg"] == "":
                master_dict[k]["kg"] = kg
    return master_dict

def update_excel_master_file(records, master_dict, output_path="Dokumen_Rekap_NoGud_Barkot_Kg.xlsx"):
    """Memperbarui atau membuat spreadsheet master Dokumen_Rekap_NoGud_Barkot_Kg.xlsx dari data Supabase."""
    print(f"[*] Memperbarui file master Excel '{output_path}'...")
    
    # Kelompokkan data per tanggal
    date_groups = {}
    for r in records:
        tgl = r.get("tanggal", "Unknown")
        if tgl not in date_groups:
            date_groups[tgl] = []
        date_groups[tgl].append(r)
        
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Hapus sheet default
    
    font_hdr = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_hdr = PatternFill("solid", fgColor="1E293B")
    fill_alt = PatternFill("solid", fgColor="F8FAFC")
    border_side = Side(style="thin", color="CBD5E1")
    border_all = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Format tanggal indonesia
    month_names = {
        "01": "Januari", "02": "Februari", "03": "Maret", "04": "April",
        "05": "Mei", "06": "Juni", "07": "Juli", "08": "Agustus",
        "09": "September", "10": "Oktober", "11": "November", "12": "Desember"
    }
    
    def format_sheet_name(d_str):
        try:
            parts = d_str.split("-")
            return f"{int(parts[2])} {month_names.get(parts[1], parts[1])}"
        except:
            return d_str
            
    # 1. Sheet Semua Data (Master)
    ws_master = wb.create_sheet(title="Semua Data (Master)")
    ws_master.views.sheetView[0].showGridLines = True
    headers_master = ["No", "Tanggal", "No Gudang", "Nomor Barcode", "Grade", "Berat (Kg)", "Status", "Keterangan"]
    
    for c_idx, h in enumerate(headers_master, start=1):
        cell = ws_master.cell(4, c_idx, h)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = align_center
        cell.border = border_all
    ws_master.row_dimensions[4].height = 24
    
    sorted_no_guds = sorted([int(k) for k in master_dict.keys() if k.isdigit()] + [k for k in master_dict.keys() if not k.isdigit()])
    r_row = 5
    for idx, nogud_key in enumerate(sorted_no_guds, start=1):
        item = master_dict[str(nogud_key)]
        row_vals = [
            idx,
            item.get("ket", ""),
            item["no_gud"],
            item["barkot"],
            item["grade"],
            item["kg"],
            item["status"],
            ""
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_master.cell(r_row, c_idx, val)
            cell.alignment = align_center
            cell.border = border_all
            cell.font = Font(name="Segoe UI", size=10)
            if idx % 2 == 0:
                cell.fill = fill_alt
        ws_master.row_dimensions[r_row].height = 20
        r_row += 1
        
    for col in ws_master.columns:
        col_letter = get_column_letter(col[0].column)
        ws_master.column_dimensions[col_letter].width = 15
    ws_master.column_dimensions['A'].width = 6
    ws_master.column_dimensions['D'].width = 18

    # 2. Sheet per Tanggal
    for tgl_str in sorted(date_groups.keys()):
        sname = format_sheet_name(tgl_str)
        ws_tgl = wb.create_sheet(title=sname)
        ws_tgl.views.sheetView[0].showGridLines = True
        
        headers_tgl = ["No", "No Gudang", "Nomor Barcode", "Grade", "Berat (Kg)", "Status", "Keterangan"]
        for c_idx, h in enumerate(headers_tgl, start=1):
            cell = ws_tgl.cell(4, c_idx, h)
            cell.font = font_hdr
            cell.fill = fill_hdr
            cell.alignment = align_center
            cell.border = border_all
        ws_tgl.row_dimensions[4].height = 24
        
        tgl_records = date_groups[tgl_str]
        r_row = 5
        for idx, rec in enumerate(tgl_records, start=1):
            row_vals = [
                idx,
                rec.get("no_gud", ""),
                rec.get("barkot", "") or "",
                rec.get("grade", "") or "",
                rec.get("kg", "") if rec.get("kg") is not None else "",
                "SELESAI" if rec.get("barkot") or rec.get("is_done") else "-",
                ""
            ]
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws_tgl.cell(r_row, c_idx, val)
                cell.alignment = align_center
                cell.border = border_all
                cell.font = Font(name="Segoe UI", size=10)
                if idx % 2 == 0:
                    cell.fill = fill_alt
            ws_tgl.row_dimensions[r_row].height = 20
            r_row += 1
            
        for col in ws_tgl.columns:
            col_letter = get_column_letter(col[0].column)
            ws_tgl.column_dimensions[col_letter].width = 15
        ws_tgl.column_dimensions['A'].width = 6
        ws_tgl.column_dimensions['C'].width = 18

    try:
        wb.save(output_path)
        print(f"[OK] File master '{output_path}' berhasil diperbarui!")
    except Exception as e:
        print(f"[!] Catatan: Tidak dapat menyimpan ke '{output_path}': {e}")

def sync_all():
    """Menjalankan sinkronisasi lengkap dari Cek Barkot Supabase ke seluruh ekosistem Stock Gudang."""
    print("==================================================")
    print("SINKRONISASI CEK BARKOT (SUPABASE) -> STOCK GUDANG")
    print("==================================================")
    
    # 1. Ambil data dari Supabase
    records = fetch_supabase_records()
    master_dict = build_master_dict(records)
    
    # 2. Update Excel Master
    update_excel_master_file(records, master_dict)
    
    # 3. Update warehouse_data.json & warehouse_data.js master metadata
    print("[*] Menyinkronkan metadata master ke warehouse_data.json & warehouse_data.js...")
    if os.path.exists("warehouse_data.json"):
        with open("warehouse_data.json", "r", encoding="utf-8") as f:
            wh_json = json.load(f)
        wh_json["master"] = master_dict
        with open("warehouse_data.json", "w", encoding="utf-8") as f:
            json.dump(wh_json, f, indent=2)
            
    if os.path.exists("warehouse_data.js"):
        with open("warehouse_data.js", "r", encoding="utf-8") as f:
            content = f.read()
        js_data = f"window.WAREHOUSE_DATA = {json.dumps(wh_json, indent=2)};"
        with open("warehouse_data.js", "w", encoding="utf-8") as f:
            f.write(js_data)
    print("[OK] warehouse_data.json & warehouse_data.js berhasil disinkronkan!")
    
    # 4. Regenerasi Laporan Excel
    print("[*] Meregenerasi laporan Excel (Stock Susunan Bal)...")
    import create_report
    print("[OK] Seluruh proses sinkronisasi selesai dengan sukses!")

if __name__ == "__main__":
    sync_all()
